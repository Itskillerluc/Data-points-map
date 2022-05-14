import tkinter as tk
import tkinter.ttk as ttk
import openpyxl as exel
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg, 
NavigationToolbar2Tk)

graph1 = {
    'A': ['P'],
    'B': ['P', 'D', 'E'],
    'C': ['A', 'F', 'G'],
    'D': ['B', 'E'],
    'E': ['A', 'B', 'D'],
    'F': ['C'],
    'G': ['C']
}

Locaties = {
    'Truck':'A',
    'Koelcel':'B',
    'freezer storage':'F',
    'Productie Soft Fried':'C',
    'Weging Soft Fried':'C',
    'Finished':'D',
    'Wrapping':'E',
    'TruckEnd':'G',
    'None':''
}

padenn = {
    'AB':[['A','P'],['P','H'],['H','I'],['I','B']],
    'BC':[['B','I'],['I','H'],['H','K'],['K','N'],['N','M'],['M','L'],['L','C']],
    'DE':[['D','H'],['H','E']],
    'EF':[['E','H'],['H','I'],['I','J'],['J','F']],
    'FG':[['F','J'],['J','I'],['I','H'],['H','P'],['P','Q'],['Q','G']],
    'BA':[['A','P'],['P','H'],['H','I'],['I','B']],
    'CB':[['B','I'],['I','H'],['H','K'],['K','N'],['N','M'],['M','L'],['L','C']],
    'ED':[['D','H'],['H','E']],
    'FE':[['E','H'],['H','I'],['I','J'],['J','F']],
    'GF':[['F','J'],['J','I'],['I','H'],['H','P'],['P','Q'],['Q','G']]
}

padendirect = {
    'AB':[['A','P'],['P','B']],
    'BC':[['B','H'],['H','K'],['K','N'],['N','M'],['M','L'],['L','C']],
    'DE':[['D','H'],['H','E']],
    'EF':[['E','F']],
    'FG':[['F','Q'],['Q','G']],
    'BA':[['A','P'],['P','B']],
    'CB':[['B','H'],['H','K'],['K','N'],['N','M'],['M','L'],['L','C']],
    'ED':[['D','H'],['H','E']],
    'FE':[['E','F']],
    'GF':[['F','Q'],['Q','G']]
}

nodes = {
    'F':[300, 450],
    'B':[430,450],
    'J':[350, 490],
    'I':[500,480],
    'H':[540,480],
    'K':[555, 425],
    'D':[520, 430],
    'N':[625,420],
    'M':[625, 280],
    'L':[900, 270],
    'C':[920, 150],
    'P':[680, 480],
    'Q':[760, 500],
    'E':[630, 510],
    'A':[680, 570],
    'G':[760, 570]

}
class draw:
    def MainLoop():
        tk.mainloop()
    def main():
        root = tk.Tk()
        root.title("Path layout")
        canvas = tk.Canvas(root, width=1290, height=857)
        canvas.pack()
        circle = lambda x, y, color: canvas.create_oval(20+x, 20+y, 30+x, 30+y, fill=color)
        bg = tk.PhotoImage(file = r"Background.gif")
        canvas.create_image(0, 0,image=bg, anchor=tk.NW)

        line = lambda point1, point2, line, line2, color: canvas.create_line(nodes[point1][0]+25, nodes[point1][1]+25, nodes[point2][0]+25, nodes[point2][1]+25, fill=color,width=((Linesdirect.count(line)/50).__round__())+((Linesdirect.count(line2)/50).__round__()))

        line('A', 'P', ['A','P'],['P','A'], 'red')
        line('P', 'B', ['P','B'],['B','P'], 'red')
        line('B', 'H', ['B','H'],['H','B'], 'lime')
        line('H', 'K', ['H','K'],['K','H'], 'lime')
        line('K', 'N', ['K','N'],['N','K'], 'lime')
        line('N', 'M', ['N','M'],['M','N'], 'lime')
        line('M', 'L', ['M','L'],['L','M'], 'lime')
        line('L', 'C', ['L','C'],['C','L'], 'lime')
        line('D', 'H', ['D','H'],['H','D'], 'yellow')
        line('H', 'E', ['H','E'],['E','H'], 'yellow')
        line('E', 'F', ['E','F'],['F','E'], 'orange')
        line('F', 'Q', ['F','Q'],['Q','F'], 'cyan')
        line('Q', 'G', ['Q','G'],['G','Q'], 'cyan')
        circle(300, 450, "lime")
        canvas.create_text(300+10,450+10,text="F")
        circle(430, 450, 'lime')
        canvas.create_text(430+10,450+10,text="B")
        circle(540, 480, 'black')
        canvas.create_text(540+10,480+10,text="H")
        circle(555, 425, 'blue')
        canvas.create_text(555+10,425+10,text="K")
        circle(520, 430, 'lime')
        canvas.create_text(520+10,430+10,text="D")
        circle(625, 420, 'blue')
        canvas.create_text(625+10,420+10,text="N")
        circle(625, 280, 'blue')
        canvas.create_text(625+10,280+10,text="M")
        circle(900, 270, 'blue')
        canvas.create_text(900+10,270+10,text="L")
        circle(920, 150, 'lime')
        canvas.create_text(920+10,150+10,text="C")
        circle(680, 480, 'blue')
        canvas.create_text(680+10,480+10,text="P")
        circle(760, 500, 'blue')
        canvas.create_text(760+10,500+20,text="Q")
        circle(630, 510, 'lime')
        canvas.create_text(630+10,510+10,text="E")
        circle(680, 570, 'lime')
        canvas.create_text(680+10,570+10,text="A")
        circle(760, 570, 'lime')
        canvas.create_text(760+10,570+10,text="G")
        
        draw.second()

    def third():
        ws2 = tk.Toplevel()
        ws2.title("Datapoints (Path Use Quantity)")
        f = Figure(figsize=(5,4), dpi=100)
        canvas = FigureCanvasTkAgg(f, master=ws2)
        canvas.get_tk_widget().grid(row=1, column=3, rowspan=6)
        p = f.gca()
        p.bar(letter, frequency)
        canvas.draw()
        draw.fourth()

    def fourth():
        ws3 = tk.Toplevel()
        ws3.title("Datapoints (Path Layout)")
        f = Figure(figsize=(5,4), dpi=100)
        canvas = FigureCanvasTkAgg(f, master=ws3)
        canvas.get_tk_widget().grid(row=1, column=3, rowspan=6)
        p = f.gca()
        p.bar(letter2, frequency2)
        canvas.draw()
        draw.fifth()

    def fifth():
        ws4 = tk.Toplevel()
        ws4.title("Datapoints (Path Layout)")
        f = Figure(figsize=(5,4), dpi=100)
        canvas = FigureCanvasTkAgg(f, master=ws4)
        canvas.get_tk_widget().grid(row=1, column=3, rowspan=6)
        p = f.gca()
        colors = ['red','red','lime','lime','lime','lime','lime','lime','yellow','yellow','orange','cyan','cyan']
        alllines = ['AP', 'PB', 'BH', 'HK', 'KN', 'NM', 'ML', 'LC', 'DH', 'EH', 'EF', 'FQ', 'QG']
        calcamount = lambda line1, line2: (Linesdirect.count(line1))+(Linesdirect.count(line2))
        llllines = [[['A','P'],['P','A']],[['P','B'],['B','P']],[['B','H'],['H','B']],[['H','K'],['K','H']],[['K','N'],['N','K']],[['N','M'],['M','N']],[['M','L'],['L','M']],[['L','C'],['C','L']],[['D','H'],['H','D']],[['E','H'],['H','E']],[['E','F'],['F','E']],[['F','Q'],['Q','F']],[['Q','G'],['G','Q']]]
        lineamounts = [calcamount(x[0], x[1]) for x in llllines]
        p.bar(alllines, lineamounts, color = colors)
        canvas.draw()
        draw.sixth()

    def seven():
        ws6 = tk.Toplevel()
        ws6.title("Legend")
        text = tk.Label(ws6,text='Lime Nodes = POI/Location\nBlack Nodes = Measuring Nodes (this is for measuring the data)\nBlue Nodes = Path Nodes (these are to make it follow a path)\nA = Truck unloading\nB = Cooling storage\nC = Production soft fried\nD = Finished products\nE = Wrapping\nF = Cooling storage\nG = Truck loading', justify="left")
        text.pack(anchor='w')
        draw.eight()
        

    def sixth():
        ws5 = tk.Toplevel()
        ws5.title("Transport amount (Path Use Quantity)")
        f = Figure(figsize=(5,4), dpi=100)
        canvas = FigureCanvasTkAgg(f, master=ws5)
        canvas.get_tk_widget().grid(row=1, column=3, rowspan=6)
        p = f.gca()
        alllines = ['AP', 'PH', 'HI', 'BI', 'HK', 'KN', 'MN', 'LM', 'LC', 'DH', 'EH', 'JF','JI', 'PQ', 'QG']
        calcamount = lambda line1, line2: (Liness.count(line1))+(Liness.count(line2))
        llllines = [[['A','P'],['P','A']],[['P','H'],['H','P']],[['H','I'],['I','H']],[['B','I'],['I','B']],[['H','K'],['K','H']],[['K','N'],['N','K']],[['N','M'],['M','N']],[['M','L'],['L','M']],[['L','C'],['C','L']],[['D','H'],['H','D']],[['E','H'],['H','E']],[['J','F'],['F','J']],[['J','I'],['I','J']],[['P','Q'],['Q','P']],[['Q','G'],['G','Q']]]
        lineamounts = [calcamount(x[0], x[1]) for x in llllines]
        collors = ['red','magenta','purple','green','lime','lime','lime','lime','lime','yellow','orange','gray','black','cyan','cyan']
        p.bar(alllines, lineamounts, color = collors)
        canvas.draw()
        draw.seven()

    def eight():
        ws7 = tk.Toplevel()
        ws7.title("Data")
        tabControl = ttk.Notebook(ws7)
        calcamount = lambda line1, line2: (Liness.count(line1))+(Liness.count(line2))
        llllines = [[['A','P'],['P','A']],[['P','H'],['H','P']],[['H','I'],['I','H']],[['B','I'],['I','B']],[['H','K'],['K','H']],[['K','N'],['N','K']],[['N','M'],['M','N']],[['M','L'],['L','M']],[['L','C'],['C','L']],[['D','H'],['H','D']],[['E','H'],['H','E']],[['J','F'],['F','J']],[['J','I'],['I','J']],[['P','Q'],['Q','P']],[['Q','G'],['G','Q']]]
        lineamounts = str(dict(zip(['AP', 'PH', 'HI', 'BI', 'HK', 'KN', 'MN', 'LM', 'LC', 'DH', 'EH', 'JF','JI', 'PQ', 'QG'],[calcamount(x[0], x[1]) for x in llllines]))).replace("'", '').replace(",", '\n').replace('{',' ').strip('}')
        lines = tk.Label(tabControl, text=lineamounts, justify='left',anchor='w')
        tabControl.add(lines, text='Nodes Path Layout')

        nodeamounts = str(dict(zip(letter,frequency))).replace("'", '').replace(",", '\n').replace('{',' ').strip('}')
        nodesundirect = tk.Label(tabControl, text=nodeamounts, justify='left',anchor='w')

        calcamount2 = lambda line1, line2: (Linesdirect.count(line1))+(Linesdirect.count(line2))
        ln = [[['A','P'],['P','A']],[['P','B'],['B','P']],[['B','H'],['H','B']],[['H','K'],['K','H']],[['K','N'],['N','K']],[['N','M'],['M','N']],[['M','L'],['L','M']],[['L','C'],['C','L']],[['D','H'],['H','D']],[['E','H'],['H','E']],[['E','F'],['F','E']],[['F','Q'],['Q','F']],[['Q','G'],['G','Q']]]
        lineamounts2 = str(dict(zip(['AP', 'PB', 'BH', 'HK', 'KN', 'NM', 'ML', 'LC', 'DH', 'EH', 'EF', 'FQ', 'QG'],[calcamount2(x[0], x[1]) for x in ln]))).replace("'", '').replace(",", '\n').replace('{',' ').strip('}')
        linesdirect = tk.Label(tabControl, text=lineamounts2, justify='left',anchor='w')
        

        nodeamounts2 = str(dict(zip(letter2,frequency2))).replace("'", '').replace(",", '\n').replace('{',' ').strip('}')
        nodesdirect = tk.Label(tabControl, text=nodeamounts2, justify='left',anchor='w')

        tabControl.add(nodesdirect, text='Nodes Path Layout')
        tabControl.add(linesdirect, text='Lines Path Layout')
        tabControl.add(nodesundirect, text='Nodes Path Use Quantity')
        tabControl.add(lines, text='Lines Path Use Quantity')

        tabControl.pack(expand = 1, fill ="both")
        draw.MainLoop()

    def second():
        ws = tk.Toplevel()
        ws.title("Path Use Quantity")   
        canvas = tk.Canvas(ws, width=1290, height=857)
        canvas.pack()
        circle2 = lambda x, y, color: canvas.create_oval(20+x, 20+y, 30+x, 30+y, fill=color)
        bg2 = tk.PhotoImage(file = r"D:\Users\Dennis\Documents\Luc\python 3\Data-points-map\Layout FP plant (1).gif")
        canvas.create_image(0, 0,image=bg2, anchor=tk.NW)
        line = lambda point1, point2, line, line2, color: canvas.create_line(nodes[point1][0]+25, nodes[point1][1]+25, nodes[point2][0]+25, nodes[point2][1]+25, fill=color,width=((Liness.count(line)/50).__round__())+((Liness.count(line2)/50).__round__()))

        line('A', 'P', ['A','P'],['P','A'], 'red')
        line('P', 'H', ['P','H'],['H','P'], 'magenta')
        line('H', 'I', ['H','I'],['I','H'], 'purple')
        line('B', 'I', ['B','I'],['I','B'], 'green')
        line('H', 'K', ['H','K'],['K','H'], 'lime')
        line('K', 'N', ['K','N'],['N','K'], 'lime')
        line('N', 'M', ['N','M'],['M','N'], 'lime')
        line('M', 'L', ['M','L'],['L','M'], 'lime')
        line('L', 'C', ['L','C'],['C','L'], 'lime')
        line('D', 'H', ['D','H'],['H','D'], 'yellow')
        line('E', 'H', ['E','H'],['H','E'], 'orange')
        line('J', 'F', ['J','F'],['F','J'], 'gray')
        line('J', 'I', ['J','I'],['I','J'], 'black')
        line('P', 'Q', ['P','Q'],['Q','P'], 'cyan')
        line('Q', 'G', ['Q','G'],['G','Q'], 'cyan')

        circle2(300, 450, "lime")
        circle2(430, 450, 'lime')
        circle2(350, 490, 'black')
        circle2(500, 480, 'black')
        circle2(540, 480, 'black')
        circle2(555, 425, 'blue')
        circle2(520, 430, 'lime')
        circle2(625, 420, 'blue')
        circle2(625, 280, 'blue')
        circle2(900, 270, 'blue')
        circle2(920, 150, 'lime')
        circle2(680, 480, 'blue')
        circle2(760, 500, 'blue')
        circle2(630, 510, 'lime')
        circle2(680, 570, 'lime')
        circle2(760, 570, 'lime')

        canvas.create_text(300+10,450+10,text="F")
        canvas.create_text(430+10,450+10,text="B")
        canvas.create_text(350+10,490+10,text="J")
        canvas.create_text(500+10,480+10,text="I")
        canvas.create_text(540+10,480+10,text="H")
        canvas.create_text(555+10,425+10,text="K")
        canvas.create_text(520+10,430+10,text="D")
        canvas.create_text(625+10,420+10,text="N")
        canvas.create_text(625+10,280+10,text="M")
        canvas.create_text(900+10,270+10,text="L")
        canvas.create_text(920+10,150+10,text="C")
        canvas.create_text(680+10,480+10,text="P")
        canvas.create_text(760+10,500+20,text="Q")
        canvas.create_text(630+10,510+10,text="E")
        canvas.create_text(680+10,570+10,text="A")
        canvas.create_text(760+10,570+10,text="G")

        draw.third()

path = r"D:\Users\Dennis\Documents\Luc\python 3\Data-points-map\data.xlsx"
wb_obj = exel.load_workbook(path)
sheet_obj = wb_obj.active
initialcode = sheet_obj.cell(row=2, column=3)
data = []
for row in range(10, sheet_obj.max_row):
    rowdata = [sheet_obj.cell(row=row, column=i).value for i in range(1, sheet_obj.max_column)]
    data.append(rowdata)
datafiltered = []
for entry in data:
    datafiltered.append(entry[1])
    datafiltered.append(entry[2])
sublist = [datafiltered[x:x+2] for x in range(0, len(datafiltered), 2)]
prev=None
dataprocessed = []
for Product_ in sublist:
    current = Product_[1]
    if current == prev:
        dataprocessed[len(dataprocessed)-1][len(dataprocessed[len(dataprocessed)-1])-1] += ", "+Product_[0]
        prev = current
    else:
         dataprocessed.append([Product_[0]])
         prev = current
index = 0
for plaats in dataprocessed:
    dataprocessed[index] = str(plaats[0]).split(", ")
    dataprocessed[index].insert(0, "Truck")
    dataprocessed[index].insert(1, "Koelcel")
    if dataprocessed[index][len(dataprocessed[index])-1] == "Productie Soft Fried":
        dataprocessed[index].append("None")
        dataprocessed[index].append("Finished")
        dataprocessed[index].append("Wrapping")
        dataprocessed[index].append("Wrapping")
        dataprocessed[index].append("freezer storage")
        dataprocessed[index].append("freezer storage")
        dataprocessed[index].append("TruckEnd")
    index += 1
index = 0

for plek in dataprocessed:
    dataprocessed[index] = list(map(Locaties.get, plek))
    index += 1
index = 0
for plek2 in dataprocessed:
    dataprocessed[index] = ''.join(map(str,plek2))
    index += 1
index=0
for plek3 in dataprocessed:
    dataprocessed[index] = [plek3[index : index + 2] for index in range(0, len(plek3), 2)]
    index += 1

dataresult = []
for t in dataprocessed:
    for f in t:
        dataresult.append(f)

directepaden = []
for paadje in dataresult:
    if len(paadje) == 2 and paadje[0] != paadje[1] and paadje != 'CD':
        directepaden.append(padendirect[paadje])
paden = []
paadje2:str
for paadje2 in dataresult:
    if len(paadje2) == 2 and paadje2[0] != paadje2[1] and paadje2 != 'CD':
        paden.append(padenn[paadje2])

histogram = []
for histo in paden:
    for histog in histo:
        for histogr in histog:
            histogram.append(histogr)

histogramm=[]
histt = 0
odd = 1
for let in histogram:
    if let != 'A' and let != 'D':
        histt = histogram.count(let)/2
        histt = histt.__round__()
        if (odd/2).is_integer():
            histogramm.append(let)
    else:
        histogramm.append(let)
    odd += 1

frequency = []
letter = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'P', 'Q']
for lette in letter:
    frequency.append(histogramm.count(lette))

histogram2 = []
for histo2 in directepaden:
    for histog2 in histo2:
        for histogr2 in histog2:
            histogram2.append(histogr2)

histogramm2=[]
histt2 = 0
odd2 = 1
for let2 in histogram2:
    if let2 != 'A' and let2 != 'D':
        histt2 = histogram2.count(let2)/2
        histt2 = histt2.__round__()
        if (odd2/2).is_integer():
            histogramm2.append(let2)
    else:
        histogramm2.append(let2)
    odd2 += 1

frequency2 = []
letter2 = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'K', 'L', 'M', 'N', 'P', 'Q']
for lette2 in letter2:
    frequency2.append(histogramm2.count(lette2))


Linesdirect = []
for l in directepaden:
    for li in l:
        Linesdirect.append(li)

Liness = []
for ll in paden:
    for lii in ll:
        Liness.append(lii)
draw.main()